import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('food.db')
wb = load_workbook('food_menu.xlsx')
ws = wb['items']

conn.execute("create table if not exists food_items (restaurant_name text, food text, price int)")

for i in range(1,56):
    temp_str = "insert into food_items (restaurant_name, food, price) values ('{0}', '{1}', '{2}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from food_items")
for i in content:
    print(i)
